"""
Seed past_events.json from an Excel file.

The Excel file should match the format produced by this project:
  - One sheet per venue (sheet name = venue name)
  - Columns: Day, Date, Time, Event Name, Link

Only events whose date is strictly before today are imported.
Existing archive entries are never overwritten (dedup by key).

Usage:
    python seed_archive.py path/to/events.xlsx
    python seed_archive.py path/to/events.xlsx --all   # import regardless of date
"""

import argparse
import json
import pathlib
import sys
from datetime import date, datetime

import openpyxl


DOCS_DIR = pathlib.Path("docs")
ARCHIVE_PATH = DOCS_DIR / "past_events.json"


def main() -> int:
    parser = argparse.ArgumentParser(description="Seed past_events.json from an Excel file.")
    parser.add_argument("excel_path", type=pathlib.Path, help="Path to the Excel file")
    parser.add_argument(
        "--all",
        action="store_true",
        dest="import_all",
        help="Import all events regardless of date (not just past ones)",
    )
    args = parser.parse_args()

    if not args.excel_path.exists():
        print(f"ERROR: file not found: {args.excel_path}", file=sys.stderr)
        return 1

    DOCS_DIR.mkdir(exist_ok=True)

    try:
        archive = json.loads(ARCHIVE_PATH.read_text(encoding="utf-8")) if ARCHIVE_PATH.exists() else []
    except json.JSONDecodeError:
        archive = []

    archive_keys = {e["key"] for e in archive}
    today = date.today()

    try:
        wb = openpyxl.load_workbook(args.excel_path, data_only=True)
    except Exception as exc:
        print(f"ERROR: could not open Excel file: {exc}", file=sys.stderr)
        return 1

    added = 0
    skipped_future = 0
    skipped_dup = 0

    for sheet_name in wb.sheetnames:
        if sheet_name == "Past Events":
            continue

        ws = wb[sheet_name]
        venue = sheet_name

        # Find header row — look for "Date" in first 3 rows
        header_row = None
        col_map = {}
        for row_idx in range(1, 4):
            row_vals = [str(ws.cell(row_idx, c).value or "").strip().lower() for c in range(1, 7)]
            if "date" in row_vals:
                header_row = row_idx
                for col_idx, val in enumerate(row_vals, start=1):
                    col_map[val] = col_idx
                break

        if header_row is None:
            print(f"  WARNING: no header row found in sheet '{venue}', skipping")
            continue

        day_col  = col_map.get("day")
        date_col = col_map.get("date")
        time_col = col_map.get("time")
        name_col = col_map.get("event name") or col_map.get("event")
        link_col = col_map.get("link")

        if not all([date_col, name_col]):
            print(f"  WARNING: sheet '{venue}' missing required columns (Date, Event/Event Name), skipping")
            continue

        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_date = ws.cell(row_idx, date_col).value if date_col else None
            raw_name = ws.cell(row_idx, name_col).value if name_col else None

            if not raw_date or not raw_name:
                continue
            if str(raw_name).strip().lower() in ("no events found", "event name"):
                continue

            date_str = _normalize_date(raw_date)
            if not date_str:
                print(f"  WARNING: could not parse date '{raw_date}' in '{venue}', row {row_idx}, skipping")
                continue

            name_str = str(raw_name).strip()
            key = "|".join([venue.lower().strip(), date_str, name_str.lower()])

            if key in archive_keys:
                skipped_dup += 1
                continue

            if not args.import_all:
                ev_date = _parse_date(date_str)
                if ev_date is None or ev_date >= today:
                    skipped_future += 1
                    continue

            day_str  = str(ws.cell(row_idx, day_col).value or "").strip()  if day_col  else ""
            time_str = str(ws.cell(row_idx, time_col).value or "").strip() if time_col else ""
            link_val = ws.cell(row_idx, link_col).value                    if link_col else None
            link_str = str(link_val).strip() if link_val else ""

            archive.append({
                "key": key,
                "venue": venue,
                "day": day_str,
                "date": date_str,
                "time": time_str,
                "name": name_str,
                "link": link_str,
                "archived_on": today.isoformat(),
            })
            archive_keys.add(key)
            added += 1

    # Sort descending by date before saving
    archive.sort(key=lambda e: _parse_date(e.get("date", "")) or date.min, reverse=True)
    ARCHIVE_PATH.write_text(json.dumps(archive, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"\nDone.")
    print(f"  Added:          {added} event(s)")
    print(f"  Skipped (dup):  {skipped_dup} event(s)")
    if not args.import_all:
        print(f"  Skipped (future or today): {skipped_future} event(s)  [use --all to include these]")
    print(f"  Archive total:  {len(archive)} event(s)")
    print(f"\nRun 'python main.py' to regenerate HTML and Excel with the updated archive.")
    return 0


def _normalize_date(raw) -> str | None:
    """Convert various date representations to MM/DD/YYYY."""
    if isinstance(raw, datetime):
        return raw.strftime("%m/%d/%Y")
    if isinstance(raw, date):
        return raw.strftime("%m/%d/%Y")
    s = str(raw).strip()

    # Handle date ranges like "Jan 31 - Feb  1, 2026" — take only the first date
    if " - " in s:
        s = s.split(" - ")[0].strip()

    # Normalize: remove periods from month abbreviations, collapse multiple spaces
    import re
    s = re.sub(r"\s+", " ", s)          # collapse extra spaces
    s = re.sub(r"(\b[A-Za-z]+)\.", r"\1", s)  # "Feb." → "Feb"

    today = date.today()

    # Formats that include a year
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%B %d, %Y", "%b %d, %Y",
                "%B %d %Y", "%b %d %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%m/%d/%Y")
        except ValueError:
            continue

    # Formats without a year — try "Month Day YYYY" with current year, then previous year
    for fmt in ("%B %d", "%b %d"):
        try:
            for year in (today.year, today.year - 1):
                candidate = datetime.strptime(f"{s} {year}", f"{fmt} %Y")
                if candidate.date() <= today:
                    return candidate.strftime("%m/%d/%Y")
        except ValueError:
            continue

    return None


def _parse_date(date_str: str) -> date | None:
    try:
        return datetime.strptime(date_str.strip(), "%m/%d/%Y").date()
    except (ValueError, AttributeError):
        return None


if __name__ == "__main__":
    sys.exit(main())
