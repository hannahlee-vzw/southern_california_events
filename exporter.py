import pathlib

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from scrapers.base import VenueResult

HEADERS = ["Day", "Date", "Time", "Event Name", "Link"]
PAST_HEADERS = ["Venue", "Day", "Date", "Time", "Event Name", "Link"]
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LINK_FONT = Font(color="0563C1", underline="single")


def write(results: list[VenueResult], output_path: pathlib.Path, past_events: list[dict] | None = None) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove the default empty sheet

    for venue in results:
        ws = wb.create_sheet(title=venue.venue_name[:31])  # Excel sheet name max 31 chars
        _write_header(ws)
        for row_idx, event in enumerate(venue.events, start=2):
            ws.cell(row_idx, 1, event.day)
            ws.cell(row_idx, 2, event.date)
            ws.cell(row_idx, 3, event.time)
            ws.cell(row_idx, 4, event.name)
            link_cell = ws.cell(row_idx, 5, event.link)
            link_cell.hyperlink = event.link
            link_cell.font = LINK_FONT

        if not venue.events:
            ws.cell(2, 1, "No events found")

        _autofit_columns(ws)
        ws.freeze_panes = "A2"

    if past_events:
        ws = wb.create_sheet(title="Past Events")
        _write_past_header(ws)
        for row_idx, ev in enumerate(past_events, start=2):
            ws.cell(row_idx, 1, ev["venue"])
            ws.cell(row_idx, 2, ev["day"])
            ws.cell(row_idx, 3, ev["date"])
            ws.cell(row_idx, 4, ev["time"])
            ws.cell(row_idx, 5, ev["name"])
            link_cell = ws.cell(row_idx, 6, ev["link"])
            link_cell.hyperlink = ev["link"]
            link_cell.font = LINK_FONT
        _autofit_columns(ws)
        ws.freeze_panes = "A2"

    wb.save(output_path)


def _write_header(ws) -> None:
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(1, col, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _write_past_header(ws) -> None:
    for col, header in enumerate(PAST_HEADERS, start=1):
        cell = ws.cell(1, col, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _autofit_columns(ws) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 60)
